# src/excel_manager.py - Fixed with timezone support

from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment 

class ExcelManager:
    def __init__(self, excel_path="./data/LeetCodeReviewSheet.xlsx"):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.worksheet = None
        self.headers = ["LeetCode Problem", "Date Last Solved", "Next Review Date", 
                        "Number of Successful Reviews", "Status"]
    
    def _clean_datetime(self, dt):
        """Remove timezone info from datetime for Excel compatibility."""
        if dt is None:
            return None
        if isinstance(dt, datetime):
            if dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
        return dt
    
    def setup_excel(self):
        """Create or load Excel file."""
        if self.excel_path.exists():
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook.active
            return
        
        # Create new file
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "LeetCode Review"
        
        # Add headers
        for col, header in enumerate(self.headers, start=1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Set column widths
        self.worksheet.column_dimensions['A'].width = 35
        self.worksheet.column_dimensions['B'].width = 18
        self.worksheet.column_dimensions['C'].width = 18
        self.worksheet.column_dimensions['D'].width = 25
        self.worksheet.column_dimensions['E'].width = 15
        
        self.workbook.save(self.excel_path)
    
    def load_problems(self):
        """Load all problems from Excel."""
        if not self.excel_path.exists():
            return {}
        
        if self.workbook is None:
            self.setup_excel()
        
        problems = {}
        for row in range(2, self.worksheet.max_row + 1):
            name = self.worksheet.cell(row=row, column=1).value
            if name:
                problems[name] = {
                    'row': row,
                    'date_last_solved': self.worksheet.cell(row=row, column=2).value,
                    'next_review_date': self.worksheet.cell(row=row, column=3).value,
                    'successful_reviews': self.worksheet.cell(row=row, column=4).value or 0,
                    'status': self.worksheet.cell(row=row, column=5).value or "📝 New"
                }
        return problems
    
    def add_or_update_problem(self, problem_name, solve_date):
        """Add a new problem or update an existing one."""
        # ✅ Clean the date (remove timezone)
        solve_date = self._clean_datetime(solve_date)
        
        if self.workbook is None:
            self.setup_excel()
        
        # Check if problem exists
        problem_exists = False
        for row in range(2, self.worksheet.max_row + 1):
            if self.worksheet.cell(row=row, column=1).value == problem_name:
                problem_exists = True
                self.worksheet.cell(row=row, column=2, value=solve_date)
                print(f"  🔄 Updated: {problem_name}")
                break
        
        if not problem_exists:
            new_row = self.worksheet.max_row + 1
            self.worksheet.cell(row=new_row, column=1, value=problem_name)
            self.worksheet.cell(row=new_row, column=2, value=solve_date)
            next_review = solve_date + timedelta(days=7) if solve_date else datetime.now()
            # ✅ Clean the next review date too
            next_review = self._clean_datetime(next_review)
            self.worksheet.cell(row=new_row, column=3, value=next_review)
            self.worksheet.cell(row=new_row, column=4, value=0)
            self.worksheet.cell(row=new_row, column=5, value="📝 New")
            print(f"  ➕ Added: {problem_name}")
    
    def save(self):
        """Save the workbook."""
        if self.workbook:
            self.workbook.save(self.excel_path)
            print(f"💾 Saved: {self.excel_path}")
    
    def get_due_reviews(self):
        """Get problems due for review."""
        if self.workbook is None:
            self.setup_excel()
        
        problems = self.load_problems()
        today = datetime.now().date()
        due = []
        for name, data in problems.items():
            next_review = data['next_review_date']
            if isinstance(next_review, datetime):
                next_review = next_review.date()
            if next_review and next_review <= today:
                due.append({
                    'name': name,
                    'row': data['row'],
                    'next_review': next_review,
                    'successful_reviews': data['successful_reviews']
                })
        return due
    
    def update_review_result(self, problem_name, success=True):
        """Update review result."""
        if self.workbook is None:
            self.setup_excel()
        
        problems = self.load_problems()
        if problem_name not in problems:
            return False
        
        row = problems[problem_name]['row']
        current = self.worksheet.cell(row=row, column=4).value or 0
        
        if success:
            new_count = current + 1
            self.worksheet.cell(row=row, column=4, value=new_count)
            next_interval = 7 * (new_count + 1)
            next_date = datetime.now().date() + timedelta(days=next_interval)
            self.worksheet.cell(row=row, column=3, value=next_date)
            self.worksheet.cell(row=row, column=5, value="✅ Success")
        else:
            next_date = datetime.now().date() + timedelta(days=7)
            self.worksheet.cell(row=row, column=3, value=next_date)
            self.worksheet.cell(row=row, column=5, value="🔄 Retry")
        
        self.save()
        return True
    
    def get_statistics(self):
        """Get statistics."""
        if self.workbook is None:
            self.setup_excel()
        
        problems = self.load_problems()
        status_count = {}
        for data in problems.values():
            status = data['status']
            status_count[status] = status_count.get(status, 0) + 1
        
        return {
            'total': len(problems),
            'due': len(self.get_due_reviews()),
            'status_counts': status_count
        }
    
    def clear_all(self):
        """Clear all problems from Excel."""
        if self.workbook is None:
            self.setup_excel()
        self.worksheet.delete_rows(2, self.worksheet.max_row - 1)
        self.save()